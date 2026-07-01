import json
import os
import re
from io import BytesIO
from openai import OpenAI
try:
    import pymupdf
except ModuleNotFoundError:
    import fitz as pymupdf
from pathlib import Path
import base64
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from copy import copy
from dateutil import parser

##################################
# OpenAI Interface and  Prompts  #
##################################

## Initialize OpenAI client
def client(server_url: str, model: str | None = None):
    if not server_url:
        raise ValueError("RunPod server URL is required.")

    server_url = server_url.rstrip("/")
    client = OpenAI(
        api_key="dummy",                 # vLLM ignores it by default
        base_url=f"{server_url}/v1"
    )

    models = client.models.list()
    available_models = [m.id for m in models.data]

    if model and model not in available_models:
        raise ValueError(f"Qwen connected, but model '{model}' was not found.")

    return client, False, "Qwen server connected successfully."

## Loading PO Json Schema
with open("po_schema.json", "r") as f:
    po_schema = json.load(f)

## Loading Timesheet Json Schema
with open("ts_schema.json", "r") as f:
    timesheet_schema = json.load(f)

## PO Reader Prompts
PO_ROLE_ALIGNMENT_RULE = """
Extract PO details exactly from the provided PO document.

Rules:

1. Return only valid JSON matching the schema.
2. Do not invent, estimate, infer, or assume missing values.
3. If a value is not explicitly provided and cannot be derived from an explicit PO formula, return null.
4. If information is not found in the PO, return null.
5. The weekend is Friday and Saturday.

Working Hours:
5. Extract working-hour information exactly as stated.
6. Create one working_hours item per location found in the PO.
7. For onshore_or_offshore return only:
   - onshore
   - offshore
   - not_specified

Rates:
8. Extract daily and hourly rates exactly as stated.
9. Do not calculate derived rates unless the PO explicitly defines the calculation method.
10. If multiple roles exist, create one rate record per role and location.
11. Preserve role names exactly as written in the PO.
12. If no role is specified return "NOT_SPECIFIED".

Derived Rates:
13. If only a base rate is provided and the PO explicitly defines a multiplier or formula, calculate the derived rate.
14. If the multiplier is not explicitly defined, return null.
15. Set rate_source to:
    - explicit
    - calculated
    - null
16. If a derived rate is calculated, populate calculation_note with the formula used.

PO Information:
17. Determine invoicing_type as:
    - daily_rate
    - hourly_rate
    - mixed
    - not_clear

18. Extract currency exactly as written.
19. Return dates exactly as written.
20. Return monetary values as numbers without currency symbols.
"""
PO_SYSTEM_PROMPT = f"""
You are an expert Purchase Order (PO) analysis and data extraction engine.

Your task is to extract structured information from Purchase Orders and return only data that is explicitly stated in the document or can be mathematically derived from stated rules.

{PO_ROLE_ALIGNMENT_RULE}

Final Rules:
1. Return data strictly according to the provided schema.
2. Do not invent, assume, estimate, or infer information that is not present in the PO.
3. When conflicting values are found, prioritize the most specific commercial section, schedule of rates, pricing table, or compensation schedule.
4. Ignore signatures, approvals, logos, headers, footers, and administrative text unless they contain commercial information.
5. If multiple amendments, revisions, change orders, or addendums exist, use the latest revision unless explicitly instructed otherwise.
6. Return only the final structured output. Do not include explanations, commentary, assumptions, confidence scores, markdown, or additional text.
"""

## Timesheet Reader Prompts
TIMESHEET_SYSTEM_PROMPT = f"""
You are extracting factual work-record data from a document.

The document format may vary. It may be a table, scanned form, handwritten form, signed sheet, daily log, or any other layout.

Extract only facts explicitly present in the document.
Do not calculate invoice values.
Do not classify billing categories.
Do not infer rates, normal hours, overtime, weekend hours, holidays, or invoice rules.

Identify each distinct work record or attendance record in the document.

For each record, extract:
- work date, if present
- day name, if present
- any time range, if present
- any total work duration, if present
- any travel duration, if present
- any location, site, activity, remarks, or description, if present

If a value is not clearly present, return null.
If multiple pages exist, extract records from all pages.
Preserve factual values. Do not move hours between categories.
Use YYYY-MM-DD for dates when confidently possible.
Use HH:MM 24-hour format for times when confidently possible.

Data Quality Rules:

1. Timesheets may contain dates with no associated work hours, travel hours, start time, or end time.
2. Do not assume that a date entry represents work performed.
3. If a date exists but no hours or time information are present, extract the date and return null for the missing values.
4. Do not populate work_duration_hours, travel_duration_hours, start_time, end_time, or time_range unless they are explicitly present in the document.
5. Do not carry values from previous or subsequent records into empty records.
6. Do not infer that a day was worked simply because it appears in the document.
7. Empty cells, blank rows, dashes, or missing entries should be treated as null unless the document explicitly states a value.
8. If a record contains a date but all work-related fields are blank, still create a record for that date with null values.

Return only JSON matching the schema.
Ignore signatures and stamps except where they explicitly identify the person associated with the work records."""

##################################
#       PO Reader Functions      #
##################################
def has_useful_pdf_text(text, min_chars=80):
    clean_text = "".join(char for char in text if char.isalnum())
    return len(clean_text) >= min_chars

def pdf_page_to_image_content(page, dpi):
    zoom = dpi / 72
    matrix = pymupdf.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=matrix, alpha=False)
    image_bytes = pix.tobytes("png")
    image_b64 = base64.b64encode(image_bytes).decode("utf-8")
    return {
        "type": "image_url",
        "image_url": {"url": f"data:image/png;base64,{image_b64}"},
    }

def text_content(text):
    return {
        "type": "text",
        "text": text,
    }

def json_schema_response_format(schema):
    return {
        "type": "json_schema",
        "json_schema": {
            "name": schema["name"],
            "schema": schema["schema"],
            "strict": True,
        },
    }

def parse_chat_json_response(response):
    content = response.choices[0].message.content
    if isinstance(content, list):
        content = "".join(
            part.get("text", "") if isinstance(part, dict) else str(part)
            for part in content
        )

    content = str(content).strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?\s*", "", content)
        content = re.sub(r"\s*```$", "", content)

    return json.loads(content)

def po_pdf_to_openai_content(uploaded_file, dpi=220):
    doc = None
    try:
        uploaded_file.seek(0)
        file_bytes = uploaded_file.read()

        if not file_bytes:
            raise ValueError("The uploaded PDF is empty.")

        doc = pymupdf.open(stream=file_bytes, filetype="pdf")
        content = []

        for page_number, page in enumerate(doc, start=1):
            text = page.get_text().strip()
            if has_useful_pdf_text(text):
                content.append({
                    "type": "text",
                    "text": f"Page {page_number} text:\n{text}",
                })
            else:
                content.append(pdf_page_to_image_content(page, dpi))

        if not content:
            raise ValueError("No readable pages were found in the PDF.")

        return content
    except Exception as exc:
        file_name = getattr(uploaded_file, "name", "uploaded PDF")
        raise RuntimeError(f"Failed to read PO PDF '{file_name}': {exc}") from exc
    finally:
        if doc is not None:
            doc.close()


## PO Details Extraction using OpenAI
def generate_po_details(po_content, ai_client, model):
    if not po_content:
        raise ValueError("No PO content was provided for extraction.")

    user_content = [text_content(
        "Customer PO details. Some pages may be scanned images. Extract all PO and rate information from the text and images provided."
    )]
    user_content.extend(po_content)

    try:
        response = ai_client.chat.completions.create(
            model=model,
            messages=[
                {
                    "role": "system",
                    "content": PO_SYSTEM_PROMPT
                },
                {
                    "role": "user",
                    "content": user_content
                }
            ],
            response_format=json_schema_response_format(po_schema),
        )
        return parse_chat_json_response(response)
    except json.JSONDecodeError as exc:
        raise RuntimeError("PO extraction returned invalid JSON.") from exc
    except Exception as exc:
        raise RuntimeError(f"PO extraction failed: {exc}") from exc

## PO Details Extraction from original or scanned PDF
def get_po_data(po_w_pl, po_wo_pl, pl, haspl, ai_client, model):
  po_content = []

  try:
    if haspl:
      if po_w_pl is None:
        raise ValueError("PO file is required.")
      # Scenario 1: the PO PDF already includes the price list.
      po_content.extend(po_pdf_to_openai_content(po_w_pl))
    else:
      if po_wo_pl is None or pl is None:
        raise ValueError("Both the PO file and price-list file are required.")
      # Scenario 2: the PO PDF and price-list PDF are uploaded separately.
      po_content.extend(po_pdf_to_openai_content(po_wo_pl))
      po_content.extend(po_pdf_to_openai_content(pl))

    po_json = generate_po_details(po_content, ai_client, model)
    return po_json
  except Exception as exc:
    raise RuntimeError(f"Could not extract PO data: {exc}") from exc

##################################
#   Timesheet Reader Functions   #
##################################
## Timesheet Conversion to base64
def file_to_base64(uploaded_file, dpi=300):
  doc = None
  try:
    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()

    if not file_bytes:
      raise ValueError("The uploaded file is empty.")

    if uploaded_file.name.lower().endswith('.pdf'):
      doc = pymupdf.open(stream=file_bytes, filetype="pdf")
      zoom = dpi / 72
      matrix = pymupdf.Matrix(zoom, zoom)
      images = []

      for page in doc:
          pix = page.get_pixmap(
              matrix=matrix,
              alpha=False)
          image_bytes = pix.tobytes("png")
          image_b64 = base64.b64encode(image_bytes).decode("utf-8")
          images.append(image_b64)
    else:
      image_b64 = base64.b64encode(file_bytes).decode("utf-8")
      images = [image_b64]

    if not images:
      raise ValueError("No pages or images were found in the uploaded file.")

    return images
  except Exception as exc:
    file_name = getattr(uploaded_file, "name", "uploaded file")
    raise RuntimeError(f"Failed to convert timesheet '{file_name}': {exc}") from exc
  finally:
    if doc is not None:
      doc.close()

## Timesheet Details Extraction using OpenAI
def generate_ts_details(timesheet, ai_client, model):
  if not timesheet:
    raise ValueError("No timesheet images were provided for extraction.")

  user_content = [text_content("Extract the timesheet data.")]
  for image in timesheet:
    user_content.append({
        "type": "image_url",
        "image_url": {"url": f"data:image/png;base64,{image}"},
    })

  try:
    response = ai_client.chat.completions.create(
        model=model,
        messages=[
            {
                "role": "system",
                "content": TIMESHEET_SYSTEM_PROMPT
            },
            {
                "role": "user",
                "content": user_content
            }],
        response_format=json_schema_response_format(timesheet_schema),
    )

    return parse_chat_json_response(response)
  except json.JSONDecodeError as exc:
    raise RuntimeError("Timesheet extraction returned invalid JSON.") from exc
  except Exception as exc:
    raise RuntimeError(f"Timesheet extraction failed: {exc}") from exc

def normalize_day_name(day_name, date_text=None):
    day_map = {
        "SUNDAY": "SUN",
        "SUN": "SUN",
        "MONDAY": "MON",
        "MON": "MON",
        "TUESDAY": "TUE",
        "TUE": "TUE",
        "TUES": "TUE",
        "WEDNESDAY": "WED",
        "WED": "WED",
        "THURSDAY": "THUR",
        "THU": "THUR",
        "THUR": "THUR",
        "THURS": "THUR",
        "FRIDAY": "FRI",
        "FRI": "FRI",
        "SATURDAY": "SAT",
        "SAT": "SAT",
    }

    if day_name:
        cleaned_day = str(day_name).strip().upper().replace(".", "")
        if cleaned_day in day_map:
            return day_map[cleaned_day]

    if date_text:
        try:
            return clean_date(date_text).strftime("%a").upper().replace("THU", "THUR")
        except Exception:
            return None

    return None

def first_present(*values):
    for value in values:
        if value is not None:
            return value
    return None

def normalize_timesheet_records(ts_json):
    if not isinstance(ts_json, dict):
        raise ValueError("Timesheet extraction result is not a valid object.")

    records = ts_json.get("records", ts_json.get("entries", []))
    if records is None:
        records = []
    if not isinstance(records, list):
        raise ValueError("Timesheet records must be a list.")

    entries = []

    for record in records:
        if not isinstance(record, dict):
            continue
        date_text = record.get("date")
        day_name = normalize_day_name(record.get("day_name"), date_text)
        entries.append({
            "day_name": day_name,
            "date": date_text,
            "hours_on_site": first_present(
                record.get("work_duration_hours"),
                record.get("hours_on_site"),
                record.get("total_hours"),
            ),
            "from_time": record.get("start_time", record.get("from_time")),
            "to_time": record.get("end_time", record.get("to_time")),
            "travel_hours": record.get("travel_duration_hours", record.get("travel_hours")),
            "friday_hours": None,
            "saturday_hours": None,
            "time_range": record.get("time_range"),
            "location": record.get("location"),
            "description": record.get("description"),
            "raw_text": record.get("raw_text"),
        })

    return entries

## Timesheets Detail Extraction
def get_timesheets_data(uploaded_files, client, model):
    ts_details = {}

    for uploaded_file in uploaded_files:
        print(uploaded_file.name)
        try:
            timesheet = file_to_base64(uploaded_file)
            ts_json = generate_ts_details(timesheet, client, model)
        except Exception as exc:
            raise RuntimeError(f"Could not process timesheet '{uploaded_file.name}': {exc}") from exc

        pro_info = {
            "project_name": ts_json["client"],
            "order_number": ts_json["client_order_number"],
        }
        eng_name = ts_json["engineer_name"]
        entries = normalize_timesheet_records(ts_json)

        if "pro_info" not in ts_details:
            ts_details["pro_info"] = pro_info

        if eng_name not in ts_details:
            ts_details[eng_name] = {"ts1": entries}
        else:
            timesheet_count = len(ts_details[eng_name].keys()) + 1
            ts_details[eng_name][f"ts_{timesheet_count}"] = entries

    return ts_details

##################################
#      Time Classification       #
##################################
ENGINEER_BLOCKS = [
    {"name_cell": "B4", "cols": {"normal": "B", "ot": "C", "fri": "D", "sat": "E"}},
    {"name_cell": "F4", "cols": {"normal": "F", "ot": "G", "fri": "H", "sat": "I"}},
    {"name_cell": "J4", "cols": {"normal": "J", "ot": "K", "fri": "L", "sat": "M"}},
    {"name_cell": "N4", "cols": {"normal": "N", "ot": "O", "fri": "P", "sat": "Q"}},
]
DATE_START_ROW = 6
TEMPLATE_DATE_END_ROW = 36
TEMPLATE_SUMMARY_START_ROW = 37

## Supporting Functions for Time Classification
def to_num(value):
    if value in [None, "", "null"]:
        return 0
    if isinstance(value, (int, float)):
        return value
    value = str(value).strip()
    if value.lower() in ["", "null", "none", "n/a", "na", "-", "--"]:
        return 0
    lowered_value = value.lower().replace(" ", "")
    if lowered_value.endswith(("am", "pm")):
        parsed_time = parser.parse(value)
        hour = parsed_time.hour
        minute = parsed_time.minute
        return hour + (minute / 60)
    if ":" in value:
        hour, minute = value.split(":", 1)
        return int(hour) + (int(minute[:2]) / 60)
    try:
        return float(value)
    except ValueError:
        return 0

def duration_hours(from_time, to_time):
    if not from_time or not to_time:
        return 0

    start = to_num(from_time)
    end = to_num(to_time)

    if end < start:
        end += 24

    return end - start

def clean_date(date_text, day_name=None):
    if not date_text:
        return None

    date_text = str(date_text).strip()

    # ISO format is unambiguous
    try:
        return parser.parse(date_text, yearfirst=True).date()
    except Exception:
        pass

    # If day name is available, use it to choose dayfirst/monthfirst
    if day_name:
        parsed_dayfirst = parser.parse(date_text, dayfirst=True).date()
        parsed_monthfirst = parser.parse(date_text, dayfirst=False).date()

        expected_day = normalize_day_name(day_name)

        if normalize_day_name(parsed_dayfirst.strftime("%A")) == expected_day:
            return parsed_dayfirst

        if normalize_day_name(parsed_monthfirst.strftime("%A")) == expected_day:
            return parsed_monthfirst

    # Fallback to your current regional default
    return parser.parse(date_text, dayfirst=True).date()

## Time Classification Dunction
def classify_entry(entry, minimum_normal_hours):
    if not entry.get("date"):
        return None

    day_name = normalize_day_name(entry.get("day_name"), entry.get("date"))
    if not day_name:
        return None

    hours_on_site = to_num(entry.get("hours_on_site"))
    travel_hours = to_num(entry.get("travel_hours"))
    from_time = to_num(entry.get("from_time"))
    to_time = to_num(entry.get("to_time"))
    friday_hours = 0
    saturday_hours = 0
    raw_friday_hours = to_num(entry.get("friday_hours"))
    raw_saturday_hours = to_num(entry.get("saturday_hours"))

    has_time_entry = any([
        hours_on_site > 0,
        travel_hours > 0,
        from_time > 0,
        to_time > 0,
        raw_friday_hours > 0,
        raw_saturday_hours > 0
    ])

    if not has_time_entry:
        return None

    normal_hours = 0
    overtime_hours = 0

    if day_name in ["SUN", "MON", "TUE", "WED", "THUR"]:
        if hours_on_site > 0:
            total_hours = hours_on_site + travel_hours
        else:
            total_hours = duration_hours(
                entry.get("from_time"),
                entry.get("to_time")
            ) + travel_hours

        if total_hours <= minimum_normal_hours:
            normal_hours = minimum_normal_hours
            overtime_hours = 0
        else:
            normal_hours = minimum_normal_hours
            overtime_hours = total_hours - minimum_normal_hours

    elif day_name == "FRI":
        friday_hours = raw_friday_hours

        if friday_hours == 0:
            friday_hours = hours_on_site or duration_hours(
                entry.get("from_time"),
                entry.get("to_time")
            )

    elif day_name == "SAT":
        saturday_hours = raw_saturday_hours

        if saturday_hours == 0:
            saturday_hours = hours_on_site or duration_hours(
                entry.get("from_time"),
                entry.get("to_time")
            )

    return {
        "day_name": day_name,
        "date": clean_date(entry["date"], day_name),
        "normal": normal_hours,
        "ot": overtime_hours,
        "fri": friday_hours,
        "sat": saturday_hours
    }

def classify_daily_entry(entry):
    if not entry.get("date"):
        return None

    day_name = normalize_day_name(entry.get("day_name"), entry.get("date"))
    if not day_name:
        return None
    hours_on_site = to_num(entry.get("hours_on_site"))
    travel_hours = to_num(entry.get("travel_hours"))
    from_time = to_num(entry.get("from_time"))
    to_time = to_num(entry.get("to_time"))
    friday_hours = to_num(entry.get("friday_hours"))
    saturday_hours = to_num(entry.get("saturday_hours"))

    has_site_entry = any([
        hours_on_site > 0,
        from_time > 0,
        to_time > 0,
        friday_hours > 0,
        saturday_hours > 0,
        travel_hours > 0,
    ])

    if not has_site_entry:
        return None

    return {
        "day_name": day_name,
        "date": clean_date(entry["date"], day_name),
        "normal": 1,
        "ot": 0,
        "fri": 0,
        "sat": 0
    }

def classify_mixed_entry(entry, minimum_normal_hours):
    if not entry.get("date"):
        return None

    day_name = normalize_day_name(entry.get("day_name"), entry.get("date"))
    if not day_name:
        return None

    hours_on_site = to_num(entry.get("hours_on_site"))
    travel_hours = to_num(entry.get("travel_hours"))
    from_time = to_num(entry.get("from_time"))
    to_time = to_num(entry.get("to_time"))
    friday_hours = to_num(entry.get("friday_hours"))
    saturday_hours = to_num(entry.get("saturday_hours"))

    if hours_on_site > 0:
        total_hours = hours_on_site + travel_hours
    else:
        total_hours = duration_hours(
            entry.get("from_time"),
            entry.get("to_time")
        ) + travel_hours

    if day_name == "FRI":
        total_hours = friday_hours or total_hours
    elif day_name == "SAT":
        total_hours = saturday_hours or total_hours

    if total_hours <= 0:
        return None

    extra_hours = max(total_hours - minimum_normal_hours, 0)

    return {
        "day_name": day_name,
        "date": clean_date(entry["date"], day_name),
        "normal": 1,
        "ot": extra_hours if day_name in ["SUN", "MON", "TUE", "WED", "THUR"] else 0,
        "fri": extra_hours if day_name == "FRI" else 0,
        "sat": extra_hours if day_name == "SAT" else 0
    }

def classify_entry_by_invoicing_type(entry, minimum_normal_hours, invoicing_type):
    if invoicing_type == "daily":
        return classify_daily_entry(entry)
    if invoicing_type == "mixed":
        return classify_mixed_entry(entry, minimum_normal_hours)
    return classify_entry(entry, minimum_normal_hours)

def clean_rate(value):
    if pd.isna(value):
        return 0
    return value

def find_rate_row(rate_table, po, role, location):
    exact_match = rate_table[
        (rate_table["po_number"] == po)
        & (rate_table["role_name"] == role)
        & (rate_table["onshore_or_offshore"] == location)
    ]

    if not exact_match.empty:
        return exact_match.iloc[0]

    role_text = str(role).lower()
    fuzzy_match = rate_table[
        (rate_table["po_number"] == po)
        & (rate_table["onshore_or_offshore"] == location)
        & (
            rate_table["role_name"].astype(str).str.lower().str.contains(role_text, regex=False)
            | rate_table["role_name"].astype(str).apply(
                lambda value: str(value).lower() in role_text
            )
        )
    ]

    if not fuzzy_match.empty:
        return fuzzy_match.iloc[0]

    raise IndexError(
        f"No matching rate row found for PO '{po}', role '{role}', location '{location}'."
    )

def shift_formula_rows(formula, start_row, offset):
    cell_ref_pattern = re.compile(r"(\$?[A-Z]{1,3})(\$?)(\d+)")

    def shift_match(match):
        column, absolute_row, row_text = match.groups()
        row = int(row_text)
        if row >= start_row:
            return f"{column}{absolute_row}{row + offset}"
        return match.group(0)

    return cell_ref_pattern.sub(shift_match, formula)


def expand_date_rows_if_needed(ws, date_count):
    template_date_capacity = TEMPLATE_DATE_END_ROW - DATE_START_ROW + 1
    if date_count <= template_date_capacity:
        date_end_row = TEMPLATE_DATE_END_ROW
        return {
            "date_end_row": date_end_row,
            "hours_row": 37,
            "rates_row": 38,
            "bill_row": 39,
            "total_invoice_row": 40,
        }

    extra_rows = date_count - template_date_capacity
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    ws.insert_rows(TEMPLATE_SUMMARY_START_ROW, extra_rows)

    for target_row in range(TEMPLATE_SUMMARY_START_ROW, TEMPLATE_SUMMARY_START_ROW + extra_rows):
        ws.row_dimensions[target_row].height = ws.row_dimensions[TEMPLATE_DATE_END_ROW].height
        for column_index in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=TEMPLATE_DATE_END_ROW, column=column_index)
            target_cell = ws.cell(row=target_row, column=column_index)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.value = None

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row >= TEMPLATE_SUMMARY_START_ROW:
            min_row += extra_rows
            max_row += extra_rows
        elif max_row >= TEMPLATE_SUMMARY_START_ROW:
            max_row += extra_rows

        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )

    date_end_row = TEMPLATE_DATE_END_ROW + extra_rows
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula = shift_formula_rows(
                    cell.value,
                    TEMPLATE_SUMMARY_START_ROW,
                    extra_rows,
                )
                formula = re.sub(
                    r"(\$?[A-Z]{1,3}\$?6:\$?[A-Z]{1,3}\$?)36\b",
                    rf"\g<1>{date_end_row}",
                    formula,
                )
                cell.value = formula

    return {
        "date_end_row": date_end_row,
        "hours_row": 37 + extra_rows,
        "rates_row": 38 + extra_rows,
        "bill_row": 39 + extra_rows,
        "total_invoice_row": 40 + extra_rows,
    }


# Calculation Excel Sheet Generator (This Needs to be Modified)
def fill_calculation_excel(
    ts_details,
    po_daily_rates,
    po_hourly_rates,
    po_working_hours,
    role,
    location,
    po,
    invoicing_type="hourly_rate"
):
    template_path = "Calculation.xlsx"
    if not Path(template_path).exists():
        raise FileNotFoundError(f"Calculation template was not found: {template_path}")

    try:
        if invoicing_type == "daily":
            rate_row = find_rate_row(po_daily_rates, po, role, location)
            normal_rate = clean_rate(rate_row["normal_week_day"])
            ot_rate = 0
            friday_rate = 0
            saturday_rate = 0
        elif invoicing_type == "mixed":
            daily_rate_row = find_rate_row(po_daily_rates, po, role, location)
            hourly_rate_row = find_rate_row(po_hourly_rates, po, role, location)
            normal_rate = clean_rate(daily_rate_row["normal_week_day"])
            ot_rate = clean_rate(hourly_rate_row["overtime_week_day"])
            friday_rate = clean_rate(hourly_rate_row["friday"])
            saturday_rate = clean_rate(hourly_rate_row["saturday"])
        else:
            rate_row = find_rate_row(po_hourly_rates, po, role, location)
            normal_rate = clean_rate(rate_row["normal_week_day"])
            ot_rate = clean_rate(rate_row["overtime_week_day"])
            friday_rate = clean_rate(rate_row["friday"])
            saturday_rate = clean_rate(rate_row["saturday"])

        matching_hours = po_working_hours[
            (po_working_hours["po_number"] == po)
            & (po_working_hours["onshore_or_offshore"] == location)
        ]
        if matching_hours.empty:
            raise IndexError(
                f"No working-hours row found for PO '{po}', location '{location}'."
            )
        hours_row = matching_hours.iloc[0]
    except KeyError as exc:
        raise KeyError(f"Missing required rate or working-hours column: {exc}") from exc

    minimum_normal_hours = hours_row["normal_week_day"]

    try:
        wb = load_workbook(template_path)
    except Exception as exc:
        raise RuntimeError(f"Failed to open calculation template: {exc}") from exc
    ws = wb.active

    ws["B2"] = ts_details["pro_info"]["project_name"]
    ws["B3"] = po

    all_dates = set()
    engineer_data = {}

    for engineer_name, engineer_timesheets in list(ts_details.items())[1:]:
        engineer_data[engineer_name] = {}

        for ts_name, entries in engineer_timesheets.items():
            for entry in entries:
                try:
                    classified = classify_entry_by_invoicing_type(
                        entry,
                        minimum_normal_hours,
                        invoicing_type
                    )
                except Exception as exc:
                    raise RuntimeError(
                        f"Failed to classify entry for engineer '{engineer_name}' in '{ts_name}': {exc}"
                    ) from exc
                if classified is None:
                    continue
                date = classified["date"]

                all_dates.add(date)

                if date not in engineer_data[engineer_name]:
                    engineer_data[engineer_name][date] = {
                        "normal": 0,
                        "ot": 0,
                        "fri": 0,
                        "sat": 0
                    }

                engineer_data[engineer_name][date]["normal"] += classified["normal"]
                engineer_data[engineer_name][date]["ot"] += classified["ot"]
                engineer_data[engineer_name][date]["fri"] += classified["fri"]
                engineer_data[engineer_name][date]["sat"] += classified["sat"]

    sorted_dates = sorted(all_dates)

    template_rows = expand_date_rows_if_needed(ws, len(sorted_dates))
    start_row = DATE_START_ROW
    date_end_row = template_rows["date_end_row"]
    hours_row = template_rows["hours_row"]
    rates_row = template_rows["rates_row"]
    bill_row = template_rows["bill_row"]
    total_invoice_row = template_rows["total_invoice_row"]

    for i, date in enumerate(sorted_dates):
        row = start_row + i
        ws[f"A{row}"] = date
        ws[f"A{row}"].number_format = "dd-mmm-yyyy"

    for eng_index, (engineer_name, data_by_date) in enumerate(engineer_data.items()):
        if eng_index >= len(ENGINEER_BLOCKS):
            raise ValueError("Template supports maximum 4 engineers only.")

        block = ENGINEER_BLOCKS[eng_index]
        ws[block["name_cell"]] = engineer_name

        for i, date in enumerate(sorted_dates):
            row = start_row + i
            values = data_by_date.get(date, {"normal": 0, "ot": 0, "fri": 0, "sat": 0})

            ws[f'{block["cols"]["normal"]}{row}'] = values["normal"]
            ws[f'{block["cols"]["ot"]}{row}'] = values["ot"]
            ws[f'{block["cols"]["fri"]}{row}'] = values["fri"]
            ws[f'{block["cols"]["sat"]}{row}'] = values["sat"]

        for column in block["cols"].values():
            ws[f"{column}{hours_row}"] = f"=SUM({column}{start_row}:{column}{date_end_row})"
            ws[f"{column}{bill_row}"] = f"={column}{hours_row}*{column}{rates_row}"

        ws[f'{block["cols"]["normal"]}{rates_row}'] = normal_rate
        ws[f'{block["cols"]["ot"]}{rates_row}'] = ot_rate
        ws[f'{block["cols"]["fri"]}{rates_row}'] = friday_rate
        ws[f'{block["cols"]["sat"]}{rates_row}'] = saturday_rate

    ws[f"B{total_invoice_row}"] = f"=SUM(B{bill_row}:Q{bill_row})"

    try:
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
    except Exception as exc:
        raise RuntimeError(f"Failed to create calculation Excel file: {exc}") from exc

    return excel_file

##################################
#      Order Database Handler    #
##################################
## Reading PO Database
def get_orders_data():
    po_master = pd.read_csv('Tables/po_master.csv')
    po_working_hours = pd.read_csv('Tables/po_working_hours.csv')
    po_daily_rates = pd.read_csv('Tables/po_daily_rates.csv')
    po_hourly_rates = pd.read_csv('Tables/po_hourly_rates.csv')
    return po_master, po_working_hours, po_daily_rates, po_hourly_rates
